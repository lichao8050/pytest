# *-* coding : utf-8 *-*
# @time : 2023/5/18 21:11
# @author : Mr_Li
# @file : read_excel_until.py
# Excel 读取为用例对象
from copy import copy

from openpyxl import load_workbook

class ExcelReader:
    """不是针对普通Excel操作，专门针对测试用例表的读、写"""

    def __init__(self, excel_path):
        # start_row = 5 表示从Excel表的第5行开始读取
        self.start_row = 5
        self.excel_path = excel_path
        self.excel_file = load_workbook(excel_path)
        # sheet_1 = self.excel_file.worksheets[0] 表示从第一个sheet开始
        self.sheet_1 = self.excel_file.worksheets[0]

    def get_case_count(self):
        """获取测试用例的数量：return：返回数量"""
        # 首先获取这张表总共有多少行
        rows_count = self.sheet_1.max_row
        # 有效用例必须减去表头数
        real_count = rows_count - (self.start_row - 1)
        return real_count

    def get_test_case_describe(self, rows):
        """获取本条用例的描述， rows的参数是0开始，表示第1行"""
        # self.start_row + rows 起始行+想要获取的行， 1表示接口url所在的固定列是1
        return self.sheet_1.cell(self.start_row + rows, 1).value

    def get_method(self, rows):
        """获取接口请求方式"""
        # self.start_row + rows 起始行+想要获取的行， 2表示接口url所在的固定列是2
        return self.sheet_1.cell(self.start_row + rows, 2).value

    def get_url(self, rows):
        """获取接口URL"""
        # self.start_row + rows 起始行+想要获取的行， 3表示接口url所在的固定列是3
        return self.sheet_1.cell(self.start_row + rows, 3).value

    def get_headers(self, rows):
        """获取接口请求头部"""
        # self.start_row + rows 起始行+想要获取的行， 4表示接口url所在的固定列是4
        return self.sheet_1.cell(self.start_row + rows, 4).value

    def get_request_data(self, rows):
        """获取接口请求参数"""
        # self.start_row + rows 起始行+想要获取的行， 5表示接口url所在的固定列是5
        return self.sheet_1.cell(self.start_row + rows, 5).value

    def get_start_code(self, rows):
        """获取接口响应状态码"""
        # self.start_row + rows 起始行+想要获取的行， 6表示接口url所在的固定列是6
        return self.sheet_1.cell(self.start_row + rows, 6).value

    def get_response_data(self, rows):
        """接口返回数据"""
        # self.start_row + rows 起始行+想要获取的行， 8表示接口url所在的固定列是8
        return self.sheet_1.cell(self.start_row + rows, 8).value

    def get_is_true_or_fail(self, rows):
        """接口测试结果"""
        # self.start_row + rows 起始行+想要获取的行， 9表示接口url所在的固定列是9
        return self.sheet_1.cell(self.start_row + rows, 9).value

    def save_file(self, file1):
        """保存Excel"""
        self.excel_file.save(file1)

    def close_file(self):
        """关闭Excel"""
        self.excel_file.close()

    # 设置某个单元格的值
    def set_pass_or_fail(self, rows, value, text):
        self.sheet_1.cell(self.start_row + rows, value, value=text)

    def write_value(self, row, col, value, xlrd):
        data = xlrd.open_workbook(self.file)
        data_copy = copy(data)
        sheet = data_copy.get_sheet(0)  # 取得复制文件的sheet对象
        sheet.write(row, col, value)  # 在某一单元格写入value
        data_copy.save(self.file)